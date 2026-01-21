"""
Generate a standalone offline cashier HTML page from four Excel inventory files.

The input spreadsheets should contain barcode, name, and price columns in either
Chinese or Spanish variants (see the header options below). The resulting HTML
is saved next to the script by default.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Optional, Sequence
import json

from openpyxl import load_workbook

PDF_HELPER = r"""
<script>
(function(){
  function mmToPt(v){ return v * 2.83465; }
  function escapePdf(text){
    return String(text).replace(/\\/g,"\\\\").replace(/\(/g,"\\(").replace(/\)/g,"\\)").replace(/\r?\n/g," ");
  }
  class SimplePDF {
    constructor(opts){
      this.widthMm = (opts && opts.format && opts.format[0]) || 210;
      this.heightMm = (opts && opts.format && opts.format[1]) || 297;
      this.fontSize = 10;
      this.lines = [];
    }
    text(txt, x, y){
      this.lines.push({txt:String(txt), x:Number(x || 0), y:Number(y || 0)});
    }
    getTextWidth(txt){
      return String(txt).length * (this.fontSize * 0.5);
    }
    save(name){
      const wPt = mmToPt(this.widthMm);
      const hPt = mmToPt(this.heightMm);
      const parts = [];
      parts.push("BT");
      parts.push("/F1 "+this.fontSize+" Tf");
      this.lines.forEach(line=>{
        const x = mmToPt(line.x).toFixed(2);
        const y = (hPt - mmToPt(line.y)).toFixed(2);
        parts.push(`1 0 0 1 ${x} ${y} Tm (${escapePdf(line.txt)}) Tj`);
      });
      parts.push("ET");
      const content = parts.join("\n");
      const stream = `<< /Length ${content.length} >>\nstream\n${content}\nendstream`;
      const font = "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>";
      const page = `<< /Type /Page /Parent 2 0 R /MediaBox [0 0 ${wPt.toFixed(2)} ${hPt.toFixed(2)}] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>`;
      const objects = [
        "<< /Type /Catalog /Pages 2 0 R >>",
        "<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        page,
        stream,
        font
      ];
      let pdf = "%PDF-1.4\n";
      const offsets = [];
      objects.forEach((obj, idx)=>{
        offsets.push(pdf.length);
        pdf += `${idx+1} 0 obj\n${obj}\nendobj\n`;
      });
      const xrefOffset = pdf.length;
      const xref = ["xref","0 "+(objects.length+1),"0000000000 65535 f "];
      offsets.forEach(off=>{ xref.push(off.toString().padStart(10,"0")+" 00000 n "); });
      pdf += xref.join("\n")+"\n";
      pdf += "trailer << /Size "+(objects.length+1)+" /Root 1 0 R >>\nstartxref\n"+xrefOffset+"\n%%EOF";
      const blob = new Blob([pdf], {type:"application/pdf"});
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a");
      a.href = url;
      a.download = name || "ticket.pdf";
      document.body.appendChild(a); a.click(); a.remove();
      setTimeout(()=>URL.revokeObjectURL(url), 2000);
    }
  }
  window.jspdf = { jsPDF: SimplePDF };
})();
</script>
"""

BARCODE_HELPER = r"""
<script>
(function(){
  const patterns = {
    "0":"nnnwwnwnn","1":"wnnwnnnnw","2":"nnwwnnnnw","3":"wnwwnnnnn","4":"nnnwwnnnw","5":"wnnwwnnnn","6":"nnwwwnnnn","7":"nnnwnnwnw","8":"wnnwnnwnn","9":"nnwwnnwnn",
    "A":"wnnnnwnnw","B":"nnwnnwnnw","C":"wnwnnwnnn","D":"nnnnwwnnw","E":"wnnnwwnnn","F":"nnwnwwnnn","G":"nnnnnwwnw","H":"wnnnnwwnn","I":"nnwnnwwnn","J":"nnnnwwwnn",
    "K":"wnnnnnnww","L":"nnwnnnnww","M":"wnwnnnnwn","N":"nnnnwnnww","O":"wnnnwnnwn","P":"nnwnwnnwn","Q":"nnnnnnwww","R":"wnnnnnwwn","S":"nnwnnnwwn","T":"nnnnwnwwn",
    "U":"wwnnnnnnw","V":"nwwnnnnnw","W":"wwwnnnnnn","X":"nwnnwnnnw","Y":"wwnnwnnnn","Z":"nwwnwnnnn","-":"nwnnnnwnw",".":"wwnnnnwnn"," ":"nwwnnnwnn","*":"nwnnwnwnn",
    "$":"nwnwnwnnn","/":"nwnwnnnwn","+":"nwnnnwnwn","%":"nnnwnwnwn"
  };
  function patternToBars(pattern, moduleWidth, startX){
    const bars = [];
    let x = startX;
    let drawBar = true;
    for(const ch of pattern){
      const w = ch === "w" ? moduleWidth * 3 : moduleWidth;
      if(drawBar) bars.push({x, w});
      x += w;
      drawBar = !drawBar;
    }
    return {bars, endX: x};
  }
  window.buildCode39SVG = function(data){
    const input = "*"+String(data||"").toUpperCase()+"*";
    const moduleWidth = 1.2;
    const height = 40;
    const gap = moduleWidth;
    let x = 0;
    const segments = [];
    for(const ch of input){
      const pattern = patterns[ch];
      if(!pattern) return "";
      const res = patternToBars(pattern, moduleWidth, x);
      segments.push(res);
      x = res.endX + gap;
    }
    const totalWidth = x - gap;
    let svg = `<svg xmlns="http://www.w3.org/2000/svg" width="${totalWidth}" height="${height}" viewBox="0 0 ${totalWidth} ${height}">`;
    segments.forEach(seg=>{
      seg.bars.forEach(b=>{
        svg += `<rect x="${b.x}" y="0" width="${b.w}" height="${height}" fill="#000"/>`;
      });
    });
    svg += "</svg>";
    return svg;
  };
})();
</script>
"""

# Column header options in multiple languages
BARCODE_HEADERS = ["条码", "barcode", "Código", "codigo", "Codigo", "Código de barras"]
NAME_HEADERS = ["名称", "商品名称", "Producto", "nombre", "Nombre"]
PRICE_HEADERS = ["零售价", "售价", "价格", "precio", "Precio"]

INLINE_STYLES = """
* { box-sizing: border-box; }
body { font-family: ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, \"Segoe UI\", sans-serif; }
.bg-gray-100 { background-color: #f3f4f6; }
.bg-white { background-color: #ffffff; }
.bg-blue-600 { background-color: #2563eb; color: #fff; }
.bg-red-600 { background-color: #dc2626; color: #fff; }
.bg-green-600 { background-color: #16a34a; color: #fff; }
.bg-emerald-600 { background-color: #059669; color: #fff; }
.bg-gray-50 { background-color: #f9fafb; }
.bg-gray-200 { background-color: #e5e7eb; }
.text-gray-500 { color: #6b7280; }
.text-gray-600 { color: #4b5563; }
.text-gray-400 { color: #9ca3af; }
.text-gray-700 { color: #374151; }
.text-gray-800 { color: #1f2937; }
.text-white { color: #ffffff; }
.text-xs { font-size: 0.75rem; line-height: 1rem; }
.text-sm { font-size: 0.875rem; line-height: 1.25rem; }
.text-lg { font-size: 1.125rem; line-height: 1.75rem; }
.text-xl { font-size: 1.25rem; line-height: 1.75rem; }
.text-[11px] { font-size: 11px; }
.text-[10px] { font-size: 10px; }
.font-bold { font-weight: 700; }
.font-semibold { font-weight: 600; }
.rounded { border-radius: 0.25rem; }
.shadow { box-shadow: 0 1px 2px rgba(0,0,0,0.05), 0 1px 3px rgba(0,0,0,0.1); }
.p-4 { padding: 1rem; }
.p-3 { padding: 0.75rem; }
.p-2 { padding: 0.5rem; }
.p-1 { padding: 0.25rem; }
.px-2 { padding-left: 0.5rem; padding-right: 0.5rem; }
.px-3 { padding-left: 0.75rem; padding-right: 0.75rem; }
.px-4 { padding-left: 1rem; padding-right: 1rem; }
.py-1 { padding-top: 0.25rem; padding-bottom: 0.25rem; }
.py-2 { padding-top: 0.5rem; padding-bottom: 0.5rem; }
.py-3 { padding-top: 0.75rem; padding-bottom: 0.75rem; }
.py-4 { padding-top: 1rem; padding-bottom: 1rem; }
.mt-1 { margin-top: 0.25rem; }
.mt-2 { margin-top: 0.5rem; }
.mb-2 { margin-bottom: 0.5rem; }
.ml-2 { margin-left: 0.5rem; }
.space-y-3 > :not([hidden]) ~ :not([hidden]) { margin-top: 0.75rem; }
.space-y-4 > :not([hidden]) ~ :not([hidden]) { margin-top: 1rem; }
.space-y-2 > :not([hidden]) ~ :not([hidden]) { margin-top: 0.5rem; }
.flex { display: flex; }
.flex-1 { flex: 1 1 0%; }
.flex-col { flex-direction: column; }
.flex-row { flex-direction: row; }
.flex-wrap { flex-wrap: wrap; }
.items-center { align-items: center; }
.items-start { align-items: flex-start; }
.items-end { align-items: flex-end; }
.justify-between { justify-content: space-between; }
.justify-end { justify-content: flex-end; }
.gap-2 { gap: 0.5rem; }
.gap-3 { gap: 0.75rem; }
.gap-4 { gap: 1rem; }
.w-full { width: 100%; }
.w-20 { width: 5rem; }
.w-28 { width: 7rem; }
.w-40 { width: 10rem; }
.h-2 { height: 0.5rem; }
.max-h-40 { max-height: 10rem; }
.max-h-56 { max-height: 14rem; }
.overflow-y-auto { overflow-y: auto; }
.text-right { text-align: right; }
.text-center { text-align: center; }
.grid { display: grid; }
.grid-cols-1 { grid-template-columns: repeat(1,minmax(0,1fr)); }
.grid-cols-4 { grid-template-columns: repeat(4,minmax(0,1fr)); }
.border { border: 1px solid #d1d5db; }
.border-b { border-bottom: 1px solid #e5e7eb; }
.hidden { display: none; }
button { cursor: pointer; }
button:disabled, input:disabled, textarea:disabled, select:disabled { opacity: 0.6; cursor: not-allowed; }
input, textarea, select, button { font: inherit; }
input, textarea, select { border: 1px solid #d1d5db; border-radius: 0.25rem; padding: 0.5rem 0.75rem; }
@media (min-width: 768px) {
  .md\:flex-row { flex-direction: row; }
  .md\:items-center { align-items: center; }
  .md\:items-end { align-items: flex-end; }
  .md\:justify-between { justify-content: space-between; }
  .md\:grid-cols-4 { grid-template-columns: repeat(4,minmax(0,1fr)); }
}
"""

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang=\"zh-CN\">
<head>
<meta charset=\"UTF-8\"><meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">
<title>智能收银系统 · 稳定不卡版（按你指定顺序）</title>
<style>__INLINE_STYLES__</style>
__PDF_HELPER__
__BARCODE_HELPER__
</head>
<body class=\"bg-gray-100 p-4 space-y-3\">

<!-- ① 顶部：门店选择 + 今日统计 -->
<div class=\"bg-white p-4 rounded shadow flex flex-col gap-2 md:flex-row md:items-center md:justify-between\">
  <div>
    <h1 class=\"text-xl font-bold\">🏬 Shangmei Tienda</h1>
    <p class=\"text-xs text-gray-500\">4015 / 3149 双门店 · 全库存内置 · 支持后5位扫码</p>
    <p class=\"text-xs text-gray-400\">内置商品：__COUNT__ 个（启动分段加载：不卡住页面）</p>
  </div>
  <div class=\"flex flex-col items-start gap-2 md:items-end\">
    <div class=\"flex items-center gap-2\">
      <span class=\"text-sm\">当前门店:</span>
      <select id=\"storeSelect\" class=\"border rounded px-2 py-1 text-sm\">
        <option value=\"4015\">4015</option><option value=\"3149\">3149</option>
      </select>
    </div>
    <div class=\"text-xs text-gray-600\" id=\"todayStats\">今日销售：0 笔 · $0</div>
  </div>
</div>

<!-- 加载提示条 -->
<div id=\"loadingBar\" class=\"bg-white p-3 rounded shadow text-xs flex items-center justify-between gap-3\">
  <div class=\"flex-1\">
    <div class=\"font-semibold\">正在加载库存…</div>
    <div class=\"text-gray-500\">加载完成后才能扫码/搜索（防止页面转圈卡死）。</div>
  </div>
  <div class=\"w-40\">
    <div class=\"h-2 bg-gray-200 rounded\">
      <div id=\"loadingProg\" class=\"h-2 bg-blue-600 rounded\" style=\"width:0%\"></div>
    </div>
    <div class=\"text-right text-gray-500 mt-1\"><span id=\"loadingText\">0%</span></div>
  </div>
</div>

<!-- ② 扫码 + 名称搜索 -->
<div class=\"bg-white p-4 rounded shadow space-y-4\">
  <div class=\"space-y-2\">
    <div class=\"flex flex-col md:flex-row gap-2\">
      <div class=\"flex flex-1 gap-2\">
        <input id=\"barcode\" class=\"border rounded px-3 py-2 flex-1\" placeholder=\"扫码 / 输入条码（支持后5位）\" disabled>
        <input id=\"qty\" type=\"number\" value=\"1\" min=\"1\" class=\"border rounded px-2 w-20 text-center\" disabled>
        <button id=\"addBtn\" class=\"bg-blue-600 text-white px-4 rounded disabled:opacity-50\" disabled>确认</button>
      </div>
    </div>
    <p class=\"text-[11px] text-gray-500\">后5位必须唯一匹配，否则提示“不唯一”。</p>
  </div>
  <div class=\"space-y-2\">
    <div class=\"flex gap-2\">
      <input id=\"searchName\" class=\"border rounded px-3 py-2 flex-1 text-sm\" placeholder=\"按名称搜索商品（中西文都可以）\" disabled>
      <button id=\"searchBtn\" class=\"border px-3 py-2 rounded text-sm bg-gray-50 disabled:opacity-50\" disabled>搜索</button>
      <button id=\"searchClearBtn\" class=\"border px-3 py-2 rounded text-sm\">清空</button>
    </div>
    <div id=\"searchResults\" class=\"max-h-40 overflow-y-auto text-xs text-gray-700 border rounded p-2 hidden\"></div>
  </div>
</div>

<!-- ③ 购物车（收银区） -->
<div class=\"bg-white p-4 rounded shadow\">
  <h2 class=\"font-semibold mb-2\">🛒 购物车</h2>
  <div id=\"cart\" class=\"text-sm text-gray-800\"></div>
  <p class=\"text-right font-bold text-lg mt-2\">合计: $<span id=\"total\">0</span></p>
  <div class=\"flex flex-wrap gap-2 mt-2\">
    <button onclick=\"pay('现金')\" class=\"bg-green-600 text-white px-4 py-2 rounded text-sm\">现金</button>
    <button onclick=\"pay('刷卡')\" class=\"border px-4 py-2 rounded text-sm\">刷卡</button>
    <button onclick=\"clearCart()\" class=\"bg-red-600 text-white px-4 py-2 rounded text-sm\">清空</button>
    <button onclick=\"printTicket()\" class=\"border px-4 py-2 rounded text-sm\">打印小票</button>
  </div>
</div>

<!-- ④ 新商品导入（放在购物车下面） -->
<div class=\"bg-white p-4 rounded shadow space-y-2\">
  <h2 class=\"font-semibold text-sm\">📥 新商品导入（追加）</h2>
  <p class=\"text-xs text-gray-500\">每行：<code>条码,名称,价格</code>（逗号或Tab）· 只新增不覆盖。</p>
  <textarea id=\"importBox\" class=\"w-full h-20 border rounded p-2 text-xs\" placeholder=\"789000000001, Nuevo producto, 1500\"></textarea>
  <div class=\"flex gap-2 justify-end\">
    <button id=\"importBtn\" class=\"bg-emerald-600 text-white px-3 py-1 rounded text-xs\">导入并保存</button>
    <button id=\"clearImportBtn\" class=\"border px-3 py-1 rounded text-xs\">清空输入</button>
  </div>
  <p class=\"text-xs text-gray-500\" id=\"extraInfo\"></p>
</div>

<!-- ⑤ 销售报表 -->
<div class=\"bg-white p-4 rounded shadow space-y-2\">
  <h2 class=\"font-semibold\">📊 销售报表</h2>
  <div class=\"flex flex-wrap gap-2 text-xs\">
    <span class=\"text-gray-600\">范围：</span>
    <button class=\"border px-2 py-1 rounded\" onclick=\"renderReport('today')\">今日</button>
    <button class=\"border px-2 py-1 rounded\" onclick=\"renderReport('month')\">本月</button>
    <button class=\"border px-2 py-1 rounded\" onclick=\"renderReport('all')\">全部</button>
    <button class=\"border px-2 py-1 rounded bg-gray-50\" onclick=\"exportReport()\">导出CSV</button>
    <button class=\"border px-2 py-1 rounded\" onclick=\"clearSales()\">清空本门店</button>
  </div>
  <div id=\"reportSummary\" class=\"text-xs text-gray-600\"></div>
  <div id=\"reportTable\" class=\"max-h-56 overflow-y-auto text-xs text-gray-800 border rounded p-2 bg-gray-50\"></div>
</div>

<!-- ⑥ 条码标签打印 -->
<div class=\"bg-white p-4 rounded shadow space-y-2\">
  <h2 class=\"font-semibold\">🏷️ 条码标签打印</h2>
  <p class=\"text-xs text-gray-500\">名称/价格可留空，自动从库存带出。打印不再“转圈卡死”（不在新窗口加载脚本）。</p>
  <div class=\"grid grid-cols-1 md:grid-cols-4 gap-2 text-xs\">
    <input id=\"labelCode\" class=\"border rounded px-2 py-1\" placeholder=\"条码\">
    <input id=\"labelName\" class=\"border rounded px-2 py-1\" placeholder=\"名称\">
    <input id=\"labelPrice\" class=\"border rounded px-2 py-1\" placeholder=\"价格\">
    <input id=\"labelCount\" type=\"number\" value=\"1\" min=\"1\" class=\"border rounded px-2 py-1\" placeholder=\"张数\">
  </div>
  <div class=\"flex gap-2 justify-end\">
    <button onclick=\"fillLabelFromBarcode()\" class=\"border px-3 py-1 rounded text-xs\">自动填充</button>
    <button onclick=\"printBarcodeLabels()\" class=\"border px-3 py-1 rounded text-xs bg-gray-50\">生成并打印</button>
  </div>
  <p class=\"text-[10px] text-gray-500\">提示：若浏览器拦截弹窗，请允许本页面弹窗。</p>
</div>

<script>
// ====== 库存数据（数组形式，分段建索引，避免 JSON.parse 大对象卡住） ======
const BASE_DATA = __DATA__; // [code, name, price][]

// 追加导入商品
const EXTRA_KEY="sm_extra_catalog";
const SALES_KEY="sm_sales_history";

function loadExtra(){ try{return JSON.parse(localStorage.getItem(EXTRA_KEY)||"{}");}catch(e){return {}; } }
function saveExtra(x){ localStorage.setItem(EXTRA_KEY, JSON.stringify(x)); }
let extraCatalog = loadExtra(); // {code:{name,price}}

// 运行时索引（分段构建）
const baseCatalog = Object.create(null);      // code -> {name, price}
const nameIndex  = [];                        // {code, nLower, name, price}
const last5Index = Object.create(null);       // last5 -> code | ""(不唯一)

let isReady = false;

function setUIReady(ok){
  ["barcode","qty","addBtn","searchName","searchBtn"].forEach(id=>{
    const el=document.getElementById(id);
    if(!el) return;
    el.disabled = !ok;
  });
}

function buildIndexesChunked(){
  let i=0;
  const total = BASE_DATA.length;
  const progEl=document.getElementById("loadingProg");
  const txtEl=document.getElementById("loadingText");
  const bar=document.getElementById("loadingBar");

  function step(){
    const start = performance.now();
    while(i < total && (performance.now() - start) < 12){
      const row = BASE_DATA[i++];
      const code = row[0];
      const name = row[1];
      const price = row[2];
      baseCatalog[code] = { name, price };
      nameIndex.push({ code, nLower: String(name).toLowerCase(), name, price });
      const k5 = code.slice(-5);
      if(last5Index[k5] === undefined) last5Index[k5] = code;
      else if(last5Index[k5] !== code) last5Index[k5] = ""; // 标记不唯一
    }
    const pct = Math.floor((i/total)*100);
    progEl.style.width = pct + "%";
    txtEl.textContent = pct + "%";
    if(i < total) requestAnimationFrame(step);
    else {
      isReady = true;
      setUIReady(true);
      bar.classList.add("hidden");
      renderCart();
      updateTodayStats();
      renderReport("today");
      document.getElementById("barcode").focus();
    }
  }
  requestAnimationFrame(step);
}

// ====== 公用 ======
const cart=[];
function store(){return document.getElementById("storeSelect")?.value || "4015";}
function item(code){ return baseCatalog[code] || extraCatalog[code] || null; }

function findCode(input){
  if(baseCatalog[input] || extraCatalog[input]) return input;
  if(input.length <= 5){
    const hit = last5Index[input];
    if(hit === undefined) return null;
    if(hit === "") return null; // 不唯一
    // extra 里也可能有同last5冲突：再校验一次
    for(const k in extraCatalog){
      if(k.endsWith(input) && k !== hit) return null;
    }
    return hit;
  }
  return null;
}

function renderCart(){
  const c=document.getElementById("cart");
  c.innerHTML="";
  let sum=0;
  if(!cart.length) c.innerHTML="<p class='text-gray-400 text-xs'>购物车为空</p>";
  else cart.forEach((i,idx)=>{
    sum+=i.qty*i.price;
    c.innerHTML+=`<div class="flex justify-between items-center border-b py-1 text-xs">
      <div class="flex-1"><div>${i.name}</div><div class="text-gray-400">条码:${i.code}</div></div>
      <div class="text-right w-28"><div>x${i.qty}</div><div>$${i.qty*i.price}</div></div>
      <button class="text-red-500 text-xs ml-2" onclick="removeItem(${idx})">✕</button>
    </div>`;
  });
  document.getElementById("total").innerText=sum.toLocaleString("es-AR");
}
function removeItem(idx){cart.splice(idx,1);renderCart();}
function clearCart(){cart.length=0;renderCart();}

function addCurrent(){
  if(!isReady) return;
  const raw=(document.getElementById("barcode").value||"").trim();
  if(!raw) return;
  const code=findCode(raw);
  if(!code){alert("未识别条码（后5位不唯一/不存在）");return;}
  const it=item(code);
  if(!it){alert("未找到商品信息");return;}
  const qty=Math.max(1,parseInt(document.getElementById("qty").value||"1"));
  const ex=cart.find(x=>x.code===code);
  if(ex) ex.qty+=qty; else cart.push({code,name:it.name,price:it.price,qty});
  renderCart();
  document.getElementById("barcode").value="";
  document.getElementById("qty").value=1;
  document.getElementById("barcode").focus();
}

// ====== 搜索（使用 nameIndex，不扫描对象key） ======
function searchByName(){
  if(!isReady) return;
  const kw=(document.getElementById("searchName").value||"").trim().toLowerCase();
  const box=document.getElementById("searchResults");
  if(!kw){box.classList.add("hidden");box.innerHTML="";return;}
  const res=[];
  for(let i=0;i<nameIndex.length;i++){
    const r=nameIndex[i];
    if(r.nLower.includes(kw)){ res.push(r); if(res.length>=50) break; }
  }
  // extra 也搜
  if(res.length<50){
    for(const code in extraCatalog){
      const n=(extraCatalog[code].name||"").toLowerCase();
      if(n.includes(kw)){ res.push({code, name: extraCatalog[code].name, price: extraCatalog[code].price}); if(res.length>=50) break; }
    }
  }
  box.innerHTML=res.length?res.map(r=>`<div class="flex justify-between items-center border-b py-1">
    <div><div>${r.name}</div><div class="text-gray-400">条码:${r.code} · $${r.price}</div></div>
    <button class="border px-2 py-1 rounded text-xs" onclick="fillBarcode('${r.code}')">选择</button>
  </div>`).join(""):"<p class='text-gray-400'>没有找到</p>";
  box.classList.remove("hidden");
}
function fillBarcode(code){document.getElementById("barcode").value=code;document.getElementById("barcode").focus();}
function clearSearch(){document.getElementById("searchName").value="";const box=document.getElementById("searchResults");box.classList.add("hidden");box.innerHTML="";}

// ====== 新商品导入（追加） ======
function importExtra(){
  const ta=document.getElementById("importBox");
  const info=document.getElementById("extraInfo");
  const lines=ta.value.split(/\n/);
  let added=0;
  lines.forEach(line=>{
    const parts=line.split(/,|\t/).map(s=>s.trim()).filter(Boolean);
    if(parts.length>=3){
      const code=parts[0], name=parts[1]; const price=parseFloat(parts[2]);
      if(code&&name&&!isNaN(price)&&!baseCatalog[code]&&!extraCatalog[code]){extraCatalog[code]={name,price};added++;}
    }
  });
  if(added){saveExtra(extraCatalog);info.textContent="已追加 "+added+" 个新商品（本机保存）";} else info.textContent="没有有效新增（或条码已存在)";
}

// ====== 销售记录 / 报表 ======
function loadSales(){try{return JSON.parse(localStorage.getItem(SALES_KEY)||"[]");}catch(e){return [];}}
function saveSales(x){localStorage.setItem(SALES_KEY,JSON.stringify(x));}

function updateTodayStats(){
  const s=store(), list=loadSales(); const now=new Date();
  const y=now.getFullYear(),m=now.getMonth(),d=now.getDate();
  let c=0,t=0;
  list.forEach(r=>{const dt=new Date(r.time); if(r.store===s&&dt.getFullYear()===y&&dt.getMonth()===m&&dt.getDate()===d){c++;t+=r.total;}});
  document.getElementById("todayStats").innerText=`门店 ${s} · 今日销售：${c} 笔 · $${t.toLocaleString("es-AR")}`;
}

function pay(type){
  if(!cart.length) return;
  const total=cart.reduce((a,i)=>a+i.qty*i.price,0);
  const s=store(); const now=new Date();
  const {jsPDF}=window.jspdf;
  const doc=new jsPDF({unit:"mm",format:[58,200]});
  let y=6;
  const head=`Shangmei Tienda ${s}`;
  doc.text(head,(58-doc.getTextWidth(head))/2,y); y+=5;
  doc.text("TICKET / 收银小票",4,y); y+=5;
  const pad=n=>String(n).padStart(2,"0");
  const f=`${now.getFullYear()}-${pad(now.getMonth()+1)}-${pad(now.getDate())} ${pad(now.getHours())}:${pad(now.getMinutes())}`;
  doc.text("Fecha: "+f,2,y); y+=5;
  doc.text("----------------------------------",2,y); y+=5;
  cart.forEach(i=>{doc.text(i.name,2,y); y+=4; doc.text("Cod: "+i.code,2,y); y+=4; doc.text(`x${i.qty} Unit:$${i.price} Sub:$${i.qty*i.price}`,2,y); y+=5; doc.text("----------------------------------",2,y); y+=5;});
  doc.text("TOTAL: $"+total.toLocaleString("es-AR"),2,y); y+=5;
  doc.text("Pago: "+type,2,y); y+=5;
  doc.save(`ticket_${s}_${Date.now()}.pdf`);

  const sales=loadSales();
  sales.push({id:Date.now(),store:s,time:now.toISOString(),paymentType:type,total,items:cart.map(i=>({code:i.code,name:i.name,qty:i.qty,price:i.price}))});
  saveSales(sales);
  updateTodayStats();
  renderReport("today");
  clearCart();
  alert("支付完成 · 已保存销售记录");
}

function buildPrintContent(){
  const s=store(), now=new Date(); const pad=n=>String(n).padStart(2,"0");
  const f=`${now.getFullYear()}-${pad(now.getMonth()+1)}-${pad(now.getDate())} ${pad(now.getHours())}:${pad(now.getMinutes())}`;
  let html=`<div style="width:58mm;font-family:monospace;font-size:10px;"><div style="text-align:center;font-weight:bold;">Shangmei Tienda ${s}</div><div style="text-align:center;">TICKET</div><div>Fecha: ${f}</div><div>--------------------------------</div>`;
  let total=0;
  cart.forEach(i=>{const sub=i.qty*i.price; total+=sub; html+=`<div>${i.name}</div><div>Cod: ${i.code}</div><div>x${i.qty} Unit:$${i.price} Sub:$${sub}</div><div>--------------------------------</div>`;});
  html+=`<div style="text-align:right;font-weight:bold;">TOTAL: $${total}</div></div>`;
  return html;
}
function printTicket(){
  if(!cart.length){alert("购物车为空，无法打印");return;}
  const w=window.open("","_blank","width=400,height=600");
  w.document.write("<html><head><title>Ticket</title></head><body>"+buildPrintContent()+"</body></html>");
  w.document.close(); w.focus(); w.print(); w.close();
}

function filterSales(range){
  const s=store(); const all=loadSales().filter(r=>r.store===s);
  if(range==="all") return all;
  const now=new Date(); const y=now.getFullYear(),m=now.getMonth(),d=now.getDate();
  return all.filter(r=>{const t=new Date(r.time); if(range==="today") return t.getFullYear()===y&&t.getMonth()===m&&t.getDate()===d; if(range==="month") return t.getFullYear()===y&&t.getMonth()===m; return true;});
}
function renderReport(range){
  const list=filterSales(range); const sum=list.reduce((a,r)=>a+r.total,0);
  document.getElementById("reportSummary").textContent=`门店 ${store()} · 记录数：${list.length} · 合计：$${sum.toLocaleString("es-AR")}`;
  const box=document.getElementById("reportTable");
  if(!list.length){box.innerHTML="<p class='text-gray-400'>没有记录</p>";return;}
  const pad=n=>String(n).padStart(2,"0");
  box.innerHTML=list.slice().reverse().map(r=>{const t=new Date(r.time); const ds=`${t.getFullYear()}-${pad(t.getMonth()+1)}-${pad(t.getDate())} ${pad(t.getHours())}:${pad(t.getMinutes())}`; 
    return `<div class="flex justify-between border-b py-1"><div><div>${ds}</div><div class="text-gray-400 text-[10px]">方式:${r.paymentType} · 商品:${r.items.length}种</div></div><div class="font-semibold">$${r.total}</div></div>`;}).join("");
}
function exportReport(){
  const list=filterSales("all"); if(!list.length){alert("没有销售记录");return;}
  const rows=[["store","time","paymentType","total","items"].join(",")];
  list.forEach(r=>{const items=r.items.map(i=>i.code+"x"+i.qty).join("|"); rows.push([r.store,r.time,r.paymentType,r.total,items].join(","));});
  const blob=new Blob([rows.join("\n")],{type:"text/csv;charset=utf-8;"});
  const url=URL.createObjectURL(blob); const a=document.createElement("a");
  a.href=url; a.download="ventas_"+store()+".csv"; document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(url);
}
function clearSales(){
  const s=store(); const all=loadSales(); const kept=all.filter(r=>r.store!==s);
  if(confirm("确定清空门店 "+s+" 的所有销售记录吗？")){saveSales(kept); updateTodayStats(); renderReport("today");}
}

// ====== 条码标签打印：使用内置 Code39 生成SVG，再复制到新窗口（不加载外部脚本） ======
function fillLabelFromBarcode(){
  const code=(document.getElementById("labelCode").value||"").trim(); if(!code) return;
  const it=item(code); if(!it){alert("未找到该条码商品");return;}
  if(!document.getElementById("labelName").value.trim()) document.getElementById("labelName").value=it.name;
  if(!document.getElementById("labelPrice").value.trim()) document.getElementById("labelPrice").value=it.price;
}

function buildBarcodeSVG(code){
  return typeof buildCode39SVG === "function" ? buildCode39SVG(code) : "";
}

function printBarcodeLabels(){
  const code=(document.getElementById("labelCode").value||"").trim();
  let name=(document.getElementById("labelName").value||"").trim();
  let price=(document.getElementById("labelPrice").value||"").trim();
  let count=parseInt(document.getElementById("labelCount").value||"1");
  if(!code){alert("请先输入条码");return;}
  const it=item(code); if(!name&&it) name=it.name; if(!price&&it) price=it.price;
  if(!count||count<1) count=1;

  const svgHTML = buildBarcodeSVG(code);
  const w=window.open("","_blank","width=400,height=600");
  if(!w){alert("弹窗被拦截，请允许弹窗");return;}
  w.document.write(`<html><head><meta charset="UTF-8"><title>Labels</title></head>
  <body style="margin:0;padding:4px;font-family:Arial,sans-serif;">`);
  for(let i=0;i<count;i++){
    w.document.write(`<div style="width:58mm;border-bottom:1px dashed #ccc;margin-bottom:4px;padding-bottom:4px;">
      <div style="font-size:10px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">${name||""}</div>
      ${svgHTML ? svgHTML : "<div style=\"font-size:10px;\">"+code+"</div>"}
      <div style="font-size:10px;">$${price||""}</div>
    </div>`);
  }
  w.document.write(`<script>window.onload=function(){window.print();};</script></body></html>`);
  w.document.close();
}

// ====== 事件 ======
document.getElementById("addBtn").onclick=addCurrent;
document.getElementById("barcode").addEventListener("keydown",e=>{if(e.key==="Enter") addCurrent();});
document.getElementById("searchBtn").onclick=searchByName;
document.getElementById("searchClearBtn").onclick=clearSearch;
document.getElementById("searchName").addEventListener("keydown",e=>{if(e.key==="Enter") searchByName();});
document.getElementById("importBtn").onclick=importExtra;
document.getElementById("clearImportBtn").onclick=()=>{document.getElementById("importBox").value="";document.getElementById("extraInfo").textContent="";};
document.getElementById("storeSelect").addEventListener("change",()=>{updateTodayStats();renderReport("today");});

// 初始化
setUIReady(false);
renderCart();
updateTodayStats();
renderReport("today");
buildIndexesChunked();
</script>
</body></html>
"""


def find_header_index(headers: Sequence[str], options: Iterable[str]) -> Optional[int]:
    """Return the index of the first matching header option, if any."""
    for option in options:
        try:
            return headers.index(option)
        except ValueError:
            continue
    return None


def load_inventory(paths: Sequence[Path]) -> List[List[object]]:
    """Load inventory rows as [code, name, price] lists from multiple Excel files."""
    rows: List[List[object]] = []
    for path in paths:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        i_code = find_header_index(headers, BARCODE_HEADERS)
        i_name = find_header_index(headers, NAME_HEADERS)
        i_price = find_header_index(headers, PRICE_HEADERS)
        if i_code is None or i_name is None or i_price is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            code, name, price = row[i_code], row[i_name], row[i_price]
            if not code or not name or price is None:
                continue
            try:
                price_f = float(price)
            except Exception:
                continue
            code_s = str(code).strip()
            name_s = str(name).strip()
            rows.append([code_s, name_s, price_f])
    return rows


def deduplicate_by_barcode(rows: Iterable[List[object]]) -> List[List[object]]:
    """Deduplicate rows by barcode, keeping the last occurrence."""
    tmp = {}
    for code, name, price in rows:
        tmp[code] = [code, name, price]
    return list(tmp.values())


def render_html(rows: Sequence[Sequence[object]]) -> str:
    data_js = json.dumps(rows, ensure_ascii=False)
    html = HTML_TEMPLATE.replace("__COUNT__", str(len(rows)))
    html = html.replace("__DATA__", data_js)
    html = html.replace("__INLINE_STYLES__", INLINE_STYLES)
    html = html.replace("__PDF_HELPER__", PDF_HELPER)
    html = html.replace("__BARCODE_HELPER__", BARCODE_HELPER)
    return html


def main(paths: Sequence[Path], output: Path) -> Path:
    rows = load_inventory(paths)
    data = deduplicate_by_barcode(rows)
    html = render_html(data)
    output.write_text(html, encoding="utf-8")
    return output


if __name__ == "__main__":
    default_paths = [
        Path("/mnt/data/3149云知识库 1.xlsx"),
        Path("/mnt/data/3149云知识库 2.xlsx"),
        Path("/mnt/data/3149云知识库 3.xlsx"),
        Path("/mnt/data/3149云知识库 4.xlsx"),
    ]
    out = Path("/mnt/data/智能收银系统_不卡转圈_稳定版.html")
    result = main(default_paths, out)
    print(result)
